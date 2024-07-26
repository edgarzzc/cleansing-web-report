from django.shortcuts import render, redirect, get_object_or_404
from django.forms import ModelForm
import csv
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tempfile

from books.models import Book
import pandas as pd

from django.http import FileResponse, HttpResponseNotFound
import os
from django.conf import settings
from django.contrib import messages

class BookForm(ModelForm):
    class Meta:
        model = Book
        fields = ['name', 'pages']

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def book_list(request, template_name='books/book_list.html'):
    book = Book.objects.all()
    data = {}
    data['object_list'] = book
    return render(request, template_name, data)

def book_view(request, pk, template_name='books/book_detail.html'):
    book= get_object_or_404(Book, pk=pk)    
    return render(request, template_name, {'object':book})

def book_create(request, template_name='books/book_form.html'):
    context = {}
    if request.method == 'GET':
        # Clear the session variable at the beginning of a new page load
        request.session.pop('generated_file_name', None)

    if request.method == "POST":
        if 'file1' in request.FILES and 'file2' in request.FILES:
            file1 = request.FILES['file1']
            file2 = request.FILES['file2']

            # Check for invalid file format
            if file1.content_type not in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'text/csv'] or file2.content_type not in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'text/csv']:
                messages.error(request, 'Invalid file format: Please upload Excel files (.xlsx) or CSV files (.csv) only')
                request.session.pop('generated_file_name', None)
                return render(request, template_name)
            
            max_possible_columns = 100
            # Define the column names for each file
            columns_file_a = [f'col{i}' for i in range(1, max_possible_columns+1)]  # Replace W with the actual number of columns in File A
            columns_file_b = [f'col{i}' for i in range(1, max_possible_columns + 1)]  # Replace P with the actual number of columns in File B

            temp_dir = 'temp'
            if not os.path.exists(temp_dir):os.makedirs(temp_dir)

            if file1.content_type == 'text/csv':
                # # Save df1 as an Excel file
                df1 = pd.read_csv(file1, names=columns_file_a, 
                                  header=None,skip_blank_lines=False, 
                                  infer_datetime_format=True)

                # Save the Excel workbook
                temp_excel_path1 = os.path.join(settings.MEDIA_ROOT, 'temp', 
                                                'converted_file1.xlsx')
                df1.to_excel(temp_excel_path1, index=False,header=None)
                # Read the Excel file back into df1
                df1 = pd.read_excel(temp_excel_path1, header=None)
                # Delete the temporary Excel file

                # Open the saved Excel file
                wb = openpyxl.load_workbook(temp_excel_path1)
                ws = wb.active

                # Iterate over the cells and check for numeric values stored as text
                for row in ws.iter_rows(min_row=2, max_col=ws.max_column, 
                                        max_row=ws.max_row):  # skipping the header row
                    for cell in row:
                        # If the cell value is a string that can 
                        # be represented as a number
                
                        if isinstance(cell.value, str) and is_number(cell.value):
                            # Set the cell value to a float
                            cell.value = float(cell.value)

                # Save the changes to the Excel file
                wb.save(temp_excel_path1)

                df1 = pd.read_excel(temp_excel_path1, header=None)
                df1 = df1.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                os.remove(temp_excel_path1)
            else:
                df1 = pd.read_excel(file1, header=None)

            if file2.content_type == 'text/csv':
                df2 = pd.read_csv(file2, names=columns_file_b, header=None, 
                                  skip_blank_lines=False, infer_datetime_format=True)
                # Save df1 as an Excel file
                # temp_excel_path2 = 'temp_file2.xlsx'
                
                temp_excel_path2 = os.path.join(settings.MEDIA_ROOT, 'temp', 
                                                'converted_file2.xlsx')
                df2.to_excel(temp_excel_path2, index=False,header=None)
                # Read the Excel file back into df1
                df2 = pd.read_excel(temp_excel_path2, header=None)

                # Open the saved Excel file
                wb = openpyxl.load_workbook(temp_excel_path2)
                ws = wb.active

                # Iterate over the cells and check for numeric values stored as text
                for row in ws.iter_rows(min_row=2, max_col=ws.max_column, 
                                        max_row=ws.max_row):  # skipping the header row
                    for cell in row:
                        # If the cell value is a string that can be represented as a number
                        if isinstance(cell.value, str) and is_number(cell.value):
                            # Set the cell value to a float
                            cell.value = float(cell.value)

                # Save the changes to the Excel file
                wb.save(temp_excel_path2)
                
                df2 = pd.read_excel(temp_excel_path2, header=None)
                df2 = df2.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                # Delete the temporary Excel file
                os.remove(temp_excel_path2)
            else:
                df2 = pd.read_excel(file2, header=None)
            
            request.session['file1_name'] = file1.name
            request.session['file2_name'] = file2.name


            # Check conditions
            if (file1.name[2] != 'A' or file2.name[2] != 'D' or
                "Analysis" in file1.name  or "analysis" in file1.name or
                "Analysis" in file2.name  or "analysis" in file2.name):
                # Set error message in context
                messages.error(request, 'Invalid files: Please make sure you upload Anotemos File for the first file and corresponding Lesson Depict File for the second file.')
                request.session.pop('generated_file_name', None)
                return render(request, template_name)
            
            print("Value in df1 at (4, 1):", str(df1.iloc[4, 1]).strip())
            print("Value in df2 at (3, 1):", str(df2.iloc[3, 1]).strip())
            if str(df1.iloc[4, 1]).strip() != str(df2.iloc[3, 1]).strip():
                messages.error(request, 'Invalid File Pair: Please ensure that both files correspond to the same source LessonDepict file')
                request.session.pop('generated_file_name', None)
                return render(request, template_name)

            File_A_Name = file1.name[:31]
            File_D_Name = file2.name[:31]

            # Process the DataFrames and save to a new Excel file
            new_file_name = file1.name.split("2024")[0] + "2024_Cleaned_for_Analysis.xlsx"
            new_file_path = new_file_name

            with pd.ExcelWriter(new_file_name, engine='xlsxwriter') as writer:
                # Write each dataframe to a different worksheet
                df1.to_excel(writer, sheet_name= File_A_Name, index=False, header=False)
                df2.to_excel(writer, sheet_name= File_D_Name, index=False, header=False)

            print(f"New Excel file created: {new_file_name}")

            for sheet_name_to_modify in [File_A_Name, File_D_Name]:
                df = pd.read_excel(new_file_name, sheet_name=sheet_name_to_modify, header=None)
                df_modified = df.drop(df.index[0])

                # Write the modified DataFrame back to the same sheet
                with pd.ExcelWriter(new_file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_modified.to_excel(writer, sheet_name=sheet_name_to_modify, index=False, header=False)

            print(f"The first row has been removed from each sheet in: {new_file_name}")
            
            new_sheet_names = [
            "D_Background", "D_Element Report", "D_Activity Report", "D_Slide Report",
            "D_User Report", "A_Background", "A_Activity Report", "A_Commentary Report",
            "A_User Report", "A_Reply Report"
            ]

            with pd.ExcelWriter(new_file_name, engine='openpyxl', mode='a') as writer:
                # Access the workbook object of the ExcelWriter
                workbook = writer.book

                # Add new empty sheets with the specified names
                for sheet_name in new_sheet_names:
                    # Check if sheet already exists, if not, create an empty DataFrame and add it as a new sheet
                    if sheet_name not in workbook.sheetnames:
                        workbook.create_sheet(sheet_name)

            print(f"10 new empty sheets have been added to: {new_file_name}")

            source_sheet_name = File_D_Name 
            destination_sheet_name = "D_Background"

            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for the row that contains only 'Users' in the first column
            user_row_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Users')].index.min()

            # Check if 'Users' was found and its index
            if pd.isna(user_row_index):
                print("'Users' not found in the first column.")
            else:
                print(f"'Users' found at row index: {user_row_index}")

            # Copy data up to the row before 'Users'
            df_to_copy = df_source.iloc[:user_row_index] if user_row_index else df_source

            # Write the sliced DataFrame to the destination sheet
            with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

            print(f"Data has been copied to the '{destination_sheet_name}' sheet up to the row before 'Users'.")

            source_sheet_name = File_A_Name
            destination_sheet_name = "A_Background"
            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for the row that contains only 'Users' in the first column
            user_row_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Users')].index.min()

            # Check if 'Users' was found and its index
            if pd.isna(user_row_index):
                print("'Users' not found in the first column.")
            else:
                print(f"'Users' found at row index: {user_row_index}")

            # Copy data up to the row before 'Users'
            df_to_copy = df_source.iloc[:user_row_index] if user_row_index else df_source

            # Write the sliced DataFrame to the destination sheet
            with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

            print(f"Data has been copied to the '{destination_sheet_name}' sheet up to the row before 'Users'.")

            destination_sheet_name = "A_Activity Report"

            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for 'Commentary Activity'
            start_index = df_source[df_source.iloc[:, 0].astype(str).str.contains('Commentary Activity', na=False)].index.min()

            if pd.isna(start_index):
                print("'Commentary Activity' not found in the first column.")
            else:
                print(f"'Commentary Activity' found at row index: {start_index}")

                # Find the first empty line after 'Commentary Activity'
                df_sliced = df_source.iloc[start_index:].reset_index(drop=True)
                end_index = df_sliced[df_sliced.isnull().all(axis=1)].index.min()

                if pd.isna(end_index):
                    print("No empty line found after 'Commentary Activity', copying till the end.")
                    df_to_copy = df_sliced
                else:
                    print(f"Empty line found at row index: {end_index}")
                    df_to_copy = df_sliced.iloc[:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet starting from 'Commentary Activity'.")
            
            source_sheet_name = File_D_Name
            destination_sheet_name = "D_Element Report"

            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for exactly 'Elements'
            start_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Elements')].index.min()

            # Find the row index for 'Depiction Activity'
            end_index = df_source[df_source.iloc[:, 0].astype(str).str.contains('Depiction Activity', na=False)].index.min()

            if pd.isna(start_index):
                print("'Elements' not found in the first column.")
            elif pd.isna(end_index):
                print("'Depiction Activity' not found in the first column.")
            else:
                print(f"'Elements' found at row index: {start_index}")
                print(f"'Depiction Activity' found at row index: {end_index}")

                # Copy data from start_index to the row before end_index
                df_to_copy = df_source.iloc[start_index:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet from 'Elements' to just before 'Depiction Activity'.")

            destination_sheet_name = "D_Activity Report"
            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for exactly 'Depiction Activity'
            start_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Depiction Activity')].index.min()

            # Find the row index of the first NaN in the first column after 'Depiction Activity'
            end_index = df_source.iloc[start_index:].isnull().any(axis=1).idxmax() + start_index

            if pd.isna(start_index):
                print("'Depiction Activity' not found in the first column.")
            else:
                print(f"'Depiction Activity' found at row index: {start_index}")
                print(f"First NaN in the first column after 'Depiction Activity' found at row index: {end_index}")

                # Copy data from start_index to the row before end_index
                df_to_copy = df_source.iloc[start_index:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet from 'Depiction Activity' to the first NaN in the first column.")

            
            destination_sheet_name = "D_Slide Report"
            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for exactly 'Slides'
            start_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Slides')].index.min()

            if pd.isna(start_index):
                print("'Slides' not found in the first column.")
            else:
                # Find the row index of the first NaN in the first column after 'Slides'
                end_index = df_source.iloc[start_index:][df_source.iloc[start_index:, 0].isna()].index.min()

                if pd.isna(end_index):
                    print("No NaN found in the first column after 'Slides'.")
                    end_index = len(df_source)  # Default to the end of the dataframe if no NaN found
                else:
                    print(f"First NaN in the first column after 'Slides' found at row index: {end_index}")

                # Copy data from start_index to the row before end_index
                df_to_copy = df_source.iloc[start_index:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet from 'Slides' to the first NaN in the first column.")
            
            destination_sheet_name = "D_User Report"
            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for exactly 'Slides'
            start_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Users')].index.min()

            if pd.isna(start_index):
                print("'Users' not found in the first column.")
            else:
                print(f"'Users' found at row index: {start_index}")
                # Find the row index of the first NaN in the first column after 'Users'
                end_index = df_source.iloc[start_index:][df_source.iloc[start_index:, 0].isna()].index.min()

                if pd.isna(end_index):
                    print("No NaN found in the first column after 'Users'.")
                    end_index = len(df_source)  # Default to the end of the dataframe if no NaN found
                else:
                    print(f"First NaN in the first column after 'Users' found at row index: {end_index}")

                # Print the data to be copied for debugging
                df_to_copy = df_source.iloc[start_index:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet from 'Users' to the first NaN in the first column.")
            
            source_sheet_name = File_A_Name
            destination_sheet_name = "A_Commentary Report"

            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for exactly 'Slides'
            start_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Comments')].index.min()

            if pd.isna(start_index):
                print("'Comments' not found in the first column.")
            else:
                print(f"'Comments' found at row index: {start_index}")
                # Find the row index of the first NaN in the first column after 'Users'
                end_index = df_source.iloc[start_index:][df_source.iloc[start_index:, 0].isna()].index.min()

                if pd.isna(end_index):
                    print("No NaN found in the first column after 'Users'.")
                    end_index = len(df_source)  # Default to the end of the dataframe if no NaN found
                else:
                    print(f"First NaN in the first column after 'Users' found at row index: {end_index}")

                # Print the data to be copied for debugging
                df_to_copy = df_source.iloc[start_index:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet from 'Comments' to the first NaN in the first column.")
            
            destination_sheet_name = "A_User Report"

            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for exactly 'Slides'
            start_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Users')].index.min()

            if pd.isna(start_index):
                print("'Users' not found in the first column.")
            else:
                print(f"'Users' found at row index: {start_index}")
                # Find the row index of the first NaN in the first column after 'Users'
                end_index = df_source.iloc[start_index:][df_source.iloc[start_index:, 0].isna()].index.min()

                if pd.isna(end_index):
                    print("No NaN found in the first column after 'Users'.")
                    end_index = len(df_source)  # Default to the end of the dataframe if no NaN found
                else:
                    print(f"First NaN in the first column after 'Users' found at row index: {end_index}")

                # Print the data to be copied for debugging
                df_to_copy = df_source.iloc[start_index:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet from 'Users' to the first NaN in the first column.")
            
            destination_sheet_name = "A_Reply Report"
            # Read the source sheet
            df_source = pd.read_excel(new_file_path, sheet_name=source_sheet_name, header=None)

            # Find the row index for exactly 'Slides'
            start_index = df_source[df_source.iloc[:, 0].astype(str).apply(lambda x: x.strip() == 'Replies')].index.min()

            if pd.isna(start_index):
                print("'Replies' not found in the first column.")
            else:
                print(f"'Replies' found at row index: {start_index}")
                # Find the row index of the first NaN in the first column after 'Users'
                end_index = df_source.iloc[start_index:][df_source.iloc[start_index:, 0].isna()].index.min()

                if pd.isna(end_index):
                    print("No NaN found in the first column after 'Replies'.")
                    end_index = len(df_source)  # Default to the end of the dataframe if no NaN found
                else:
                    print(f"First NaN in the first column after 'Replies' found at row index: {end_index}")

                # Print the data to be copied for debugging
                df_to_copy = df_source.iloc[start_index:end_index]

                # Write the section to the destination sheet
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_to_copy.to_excel(writer, sheet_name=destination_sheet_name, index=False, header=False)

                print(f"Data has been copied to the '{destination_sheet_name}' sheet from 'Replies' to the first NaN in the first column.")

            # modify each sheet separately
            sheet_name = "D_Slide Report" 

            # Read the specified sheet into a DataFrame
            df = pd.read_excel(new_file_path, sheet_name=sheet_name, header=None)

            # Find the index where "Comment Serial No." is located
            start_index = df[df.iloc[:, 0].astype(str).str.contains("Serial No.", na=False)].index.min() + 1

            if start_index > len(df) - 1:

                if pd.isna(start_index):
                    print("'Serial No.' not found in column A.")
            else:
                # Loop through the DataFrame from start_index and append the string to each cell in column C
                for i in range(start_index, len(df)):
                    if pd.isna(df.iloc[i, 0]):
                        break  # Stop if an empty cell is encountered
                    df.iloc[i, 0] = "Slide serial no. " + str(df.iloc[i, 0])

                # Write the modified DataFrame back to the Excel file
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                print(f"String 'Slide serial no. ' has been added to column A in the sheet '{sheet_name}' from the specified start point.")

            # Load the workbook and select the sheet
            wb = load_workbook(new_file_path)
            sheet = wb[sheet_name]

            # Insert a new column after column E (which is column 5, since it's 1-indexed in openpyxl)
            sheet.insert_cols(idx=6)

            # Write "Slide image" in the second row of the new column (which is now column F)
            sheet['F2'] = "Slide image"

            # Save the workbook
            wb.save(new_file_path)

            print(f"Column inserted and 'Slide image' added in sheet '{sheet_name}'.")

            sheet_name = "D_Element Report"
            # Read the specified sheet into a DataFrame
            df = pd.read_excel(new_file_path, sheet_name=sheet_name, header=None)

            # Find the index where "Comment Serial No." is located
            start_index = df[df.iloc[:, 0].astype(str).str.contains("Serial No.", na=False)].index.min() + 1

            if start_index > len(df) - 1:

                if pd.isna(start_index):
                    print("'Serial No.' not found in column A.")
            else:
                # Loop through the DataFrame from start_index and append the string to each cell in column C
                for i in range(start_index, len(df)):
                    if pd.isna(df.iloc[i, 0]):
                        break  # Stop if an empty cell is encountered
                    df.iloc[i, 0] = "Element serial no. " + str(df.iloc[i, 0])

                # Write the modified DataFrame back to the Excel file
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                print(f"String 'Element serial no. ' has been added to column A in the sheet '{sheet_name}' from the specified start point.")
            
            sheet_name = "A_Commentary Report"
            # Read the specified sheet into a DataFrame
            df = pd.read_excel(new_file_path, sheet_name=sheet_name, header=None)

            # Find the index where "Comment Serial No." is located
            start_index = df[df.iloc[:, 0].astype(str).str.contains("Serial No.", na=False)].index.min() + 1

            if start_index > len(df) - 1:

                if pd.isna(start_index):
                    print("'Serial No.' not found in column A.")
            else:
                # Loop through the DataFrame from start_index and append the string to each cell in column C
                for i in range(start_index, len(df)):
                    if pd.isna(df.iloc[i, 0]):
                        break  # Stop if an empty cell is encountered
                    df.iloc[i, 0] = "Comment serial no. " + str(df.iloc[i, 0])

                # Write the modified DataFrame back to the Excel file
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                print(f"String 'Comment serial no. ' has been added to column A in the sheet '{sheet_name}' from the specified start point.")

            # Find the index of the first empty cell in the third column starting from the third row
            first_empty_box_index = df.iloc[2:, 2].isna().idxmax() if df.iloc[2:, 2].isna().any() else len(df)

            # Modify the third column to replace 'Slide', starting from the third row up to the first empty box
            df.iloc[2:first_empty_box_index, 2] = df.iloc[2:first_empty_box_index, 2].replace('Slide', '', regex=True).str.strip()

            # Write the modified DataFrame back to the Excel file
            with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            print(f"All occurrences of 'Slide' have been cleaned from the third column starting from the third row in the sheet '{sheet_name}'.")

            # Remove all spaces within the strings in the third column (labeled as '2'), starting from the second row
            for i in range(1, len(df)):  # Start from second row (index 1)
                if pd.notna(df.iat[i, 2]):  # Check if the cell is not NaN
                    df.iat[i, 2] = df.iat[i, 2].replace(' ', '').strip()

            # Write the modified DataFrame back to the Excel file
            with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            print(f"All spaces have been removed from the third column starting from the second row in the sheet '{sheet_name}'.")

            # Load the workbook and select the sheet
            wb = load_workbook(new_file_path)
            sheet = wb[sheet_name]

            # Insert four new columns after column C (which is the 3rd column, so new columns start at index 4)
            number_of_columns_to_insert = 4
            sheet.insert_cols(idx=4, amount=number_of_columns_to_insert)

            # Set the specific text in the second row of each new column
            headers = ["Slide Serial No.", "Slide Title", "Slide URL", "Slide Image"]
            for i, header in enumerate(headers, start=4):  # start=4 because we're starting from column D, which is the 4th column
                cell = sheet.cell(row=2, column=i)
                cell.value = header

            # Save the workbook
            wb.save(new_file_path)

            print(f"Four new columns have been inserted and labeled in the '{sheet_name}' sheet.")

            sheet.insert_cols(idx=9)

            # Set "Date" in the second row of the new column (which will be column I after insertion)
            sheet.cell(row=2, column=9).value = "Date"

            # Save the workbook
            wb.save(new_file_path)

            print(f"New column added and 'Date' set in the '{sheet_name}' sheet.")

            columns_to_delete = [18, 17, 16, 15, 14, 12, 10]

            for col_index in columns_to_delete:
                sheet.delete_cols(idx=col_index)

            # Save the workbook
            wb.save(new_file_path)

            print(f"Specified columns have been deleted from the '{sheet_name}' sheet.")
            # finish A_Commentary Report here


            sheet_name = "A_Reply Report"

            # Read the specified sheet into a DataFrame
            df = pd.read_excel(new_file_path, sheet_name=sheet_name, header=None)

            # Find the index where "Comment Serial No." is located
            start_index = df[df.iloc[:, 0].astype(str).str.contains("Serial No.", na=False)].index.min() + 1

            if start_index > len(df) - 1:

                if pd.isna(start_index):
                    print("'Serial No.' not found in column A.")
            else:
                # Loop through the DataFrame from start_index and append the string to each cell in column C
                for i in range(start_index, len(df)):
                    if pd.isna(df.iloc[i, 0]):
                        break  # Stop if an empty cell is encountered
                    df.iloc[i, 0] = "Reply serial no. " + str(df.iloc[i, 0])

                # Write the modified DataFrame back to the Excel file
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                print(f"String 'Reply serial no. ' has been added to column A in the sheet '{sheet_name}' from the specified start point.")

            # Find the index where "Comment Serial No." is located
            start_index = df[df.iloc[:, 2].astype(str).str.contains("Comment Serial No.", na=False)].index.min() + 1

            if start_index > len(df) - 1:

                if pd.isna(start_index):
                    print("'Serial No.' not found in column C.")
            else:
                # Loop through the DataFrame from start_index and append the string to each cell in column C
                for i in range(start_index, len(df)):
                    if pd.isna(df.iloc[i, 2]):
                        break  # Stop if an empty cell is encountered
                    df.iloc[i, 2] = "Comment Serial No. " + str(df.iloc[i, 2])
                # Write the modified DataFrame back to the Excel file
                with pd.ExcelWriter(new_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                print(f"String 'Comment Serial No. ' has been added to column C in the sheet '{sheet_name}' from the specified start point.")
            
            # Load the workbook and select the sheet
            wb = load_workbook(new_file_path)
            sheet = wb[sheet_name]

            # Correcting the number of columns to insert and the starting index
            number_of_columns_to_insert = 6
            # Insert six new columns right after column C (which is the 3rd column)
            sheet.insert_cols(idx=4, amount=number_of_columns_to_insert)

            # Headers that should be set in the second row of each new column
            headers = ["Slide Title", "Slide URL", "Slide image", "Comment Owner", "Comment Date", "Comment Text"]
            # Ensure the starting index matches where new columns are inserted
            for i, header in enumerate(headers, start=4):  # Corrected to start from column D, which becomes the new column after insertion
                cell = sheet.cell(row=2, column=i)
                cell.value = header

            sheet.cell(row=2, column=10).value = "Reply Owner"

            # Save the workbook
            wb.save(new_file_path)

            print(f"Six new columns have been inserted and labeled in the '{sheet_name}' sheet.")

            sheet_name = "A_Commentary Report"

            sheet = wb[sheet_name]
            last_row_with_data = 3
            for row in range(3, sheet.max_row + 1):
                if sheet[f'A{row}'].value is None:  # Check if the cell is empty
                    break
                last_row_with_data = row

            A_Commentary_Report_Last_Line = last_row_with_data+1

            sheet_D_Slide_Report = wb["D_Slide Report"]
            for row in range(3, sheet_D_Slide_Report.max_row + 1):
                if sheet_D_Slide_Report[f'A{row}'].value is None:  # Check if the cell is empty
                    break
                D_Slide_Report_Last_Line = str(row+1)

            sheet_A_Activity_Report = wb["A_Activity Report"]
            for row in range(3, sheet_A_Activity_Report.max_row + 1):
                if sheet_A_Activity_Report[f'A{row}'].value is None:  # Check if the cell is empty
                    break
                last_row_with_data_A_activity = str(row+1)

            for row1 in range(3, int(A_Commentary_Report_Last_Line) + 1):
                hyperlink_formula = f'=HYPERLINK(F{row1}, IMAGE(F{row1}))'
                sheet[f'G{row1}'] = hyperlink_formula

            # Example: Writing INDEX-MATCH formula into cells starting from D3
            # Example: Writing INDEX-MATCH formula into cells starting from D3
            formula_template = "=INDEX('D_Slide Report'!A$3:A$" + D_Slide_Report_Last_Line + ", MATCH(C{row}, 'D_Slide Report'!B$3:B$" + D_Slide_Report_Last_Line + ", 0))"
            for row in range(3, last_row_with_data+1):  # Apply until the row before the first empty cell in column A
                formula = formula_template.format(row=row)
                sheet[f'D{row}'] = formula

            # Example: Writing INDEX-MATCH formula into cells starting from D3
            formula_template = "=INDEX('D_Slide Report'!C$3:C$"+ D_Slide_Report_Last_Line + ",MATCH(C{row}, 'D_Slide Report'!B$3:B$" + D_Slide_Report_Last_Line + ", 0))"
            for row in range(3, last_row_with_data+1):  # Apply until the row before the first empty cell in column A
                formula = formula_template.format(row=row)
                sheet[f'E{row}'] = formula

            # Example: Writing INDEX-MATCH formula into cells starting from D3
            formula_template = "=INDEX('D_Slide Report'!E$3:E$" + D_Slide_Report_Last_Line + ",MATCH(C{row}, 'D_Slide Report'!B$3:B$" + D_Slide_Report_Last_Line + ",0))"
            for row in range(3, last_row_with_data+1):  # Apply until the row before the first empty cell in column A
                formula = formula_template.format(row=row)
                sheet[f'F{row}'] = formula

            # Example: Writing INDEX-MATCH formula into cells starting from D3
            formula_template = "=INDEX('A_Activity Report'!B$3:B$"+last_row_with_data_A_activity+",MATCH(A{row}, 'A_Activity Report'!D$3:D$"+last_row_with_data_A_activity+",0))"
            for row in range(3, last_row_with_data+1):  # Apply until the row before the first empty cell in column A
                formula = formula_template.format(row=row)
                sheet[f'I{row}'] = formula

            # Save the workbook with changes
            wb.save(new_file_path)

            print("Formulas have been applied to the sheet.")

            sheet_name = "D_Slide Report"
            sheet = wb[sheet_name]

            for row1 in range(3, int(D_Slide_Report_Last_Line)):
                hyperlink_formula = f'=HYPERLINK(E{row1}, IMAGE(E{row1}))'
                sheet[f'F{row1}'] = hyperlink_formula

            wb.save(new_file_path)

            sheet_name = "A_Reply Report"
            sheet = wb[sheet_name]

            if sheet['A1'].value is None:
                print("The 'A_Reply Report' sheet is empty. No formulas applied.")

            else:
                last_row_with_data = 3
                for row in range(3, sheet.max_row + 1):
                    if sheet[f'A{row}'].value is None:  # Check if the cell is empty
                        break
                    last_row_with_data = row

                A_Reply_Report_Last_Line = last_row_with_data+1
                
                last_row_with_data = 3
                for row in range(3, sheet.max_row + 1):
                    if sheet[f'A{row}'].value is None:  # Check if the cell is empty
                        break
                    last_row_with_data = row

                # Example: Writing INDEX-MATCH formula into cells starting from D3
                formula_template = "=INDEX('A_Commentary Report'!E$3:E${}, MATCH(C{{row}}, 'A_Commentary Report'!A$3:A${}, 0))".format(A_Commentary_Report_Last_Line, A_Commentary_Report_Last_Line)
                for row in range(3, last_row_with_data+1):  # Apply until the row before the first empty cell in column A
                    formula = formula_template.format(row=row)
                    sheet[f'D{row}'] = formula

                formula_template = "=INDEX('A_Commentary Report'!F$3:F${}, MATCH(C{{row}}, 'A_Commentary Report'!A$3:A${}, 0))".format(A_Commentary_Report_Last_Line, A_Commentary_Report_Last_Line)
                for row in range(3, last_row_with_data + 1):
                    formula = formula_template.format(row=row)
                    sheet[f'E{row}'] = formula

                formula_template = "=INDEX('A_Commentary Report'!H$3:H${}, MATCH(C{{row}}, 'A_Commentary Report'!A$3:A${}, 0))".format(A_Commentary_Report_Last_Line, A_Commentary_Report_Last_Line)
                for row in range(3, last_row_with_data + 1):
                    formula = formula_template.format(row=row)
                    sheet[f'G{row}'] = formula

                formula_template = "=INDEX('A_Activity Report'!B$3:B${}, MATCH(C{{row}}, 'A_Activity Report'!D$3:D${}, 0))".format(last_row_with_data_A_activity, last_row_with_data_A_activity)
                for row in range(3, last_row_with_data + 1):
                    formula = formula_template.format(row=row)
                    sheet[f'H{row}'] = formula

                formula_template = "=INDEX('A_Commentary Report'!K$3:K${}, MATCH(C{{row}}, 'A_Commentary Report'!A$3:A${}, 0))".format(A_Commentary_Report_Last_Line, A_Commentary_Report_Last_Line)
                for row in range(3, last_row_with_data + 1):
                    formula = formula_template.format(row=row)
                    sheet[f'I{row}'] = formula

                formula_template = "=INDEX('A_Activity Report'!B$3:B${}, MATCH(A{{row}}, 'A_Activity Report'!D$3:D${}, 0))".format(last_row_with_data_A_activity, last_row_with_data_A_activity)
                for row in range(3, last_row_with_data + 1):
                    formula = formula_template.format(row=row)
                    sheet[f'K{row}'] = formula
                
                for row1 in range(3, int(A_Reply_Report_Last_Line)):
                    hyperlink_formula = f'=HYPERLINK(E{row1}, IMAGE(E{row1}))'
                    sheet[f'F{row1}'] = hyperlink_formula

                wb.save(new_file_path)
                print("Formulas have been applied to the new sheet.")

                # Iterate over all sheets you want to modify
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
                        for cell in row:
                            if isinstance(cell.value, str) and "&nbsp;" in cell.value:
                                cell.value = cell.value.replace("&nbsp;", " ")
                            if isinstance(cell.value, str) and ";;" in cell.value:
                                cell.value = cell.value.replace(";;", "\n")
                            if isinstance(cell.value, str) and ";" in cell.value:
                                cell.value = cell.value.replace(";", ",")

                # Save the workbook with the replaced characters
                wb.save(new_file_path)

            # //key point here
            messages.success(request, 'Your reports have been successfully cleansed.')
            request.session['generated_file_name'] = new_file_name 
            print("Generated file name:", request.session['generated_file_name'])

    context['file1_name'] = request.session.get('file1_name')
    context['file2_name'] = request.session.get('file2_name')
    return render(request, template_name)
    
def download_excel(request):
    file_name = request.session.get('generated_file_name', 'default_file_name.xlsx')
    file_path = os.path.join(settings.BASE_DIR, file_name)

    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)
        return response
    else:
        return HttpResponseNotFound('The requested file was not found on the server.')

def book_update(request, pk, template_name='books/book_form.html'):
    book= get_object_or_404(Book, pk=pk)
    form = BookForm(request.POST or None, instance=book)
    if form.is_valid():
        form.save()
        return redirect('book_list')
    return render(request, template_name, {'form':form})

def book_delete(request, pk, template_name='books/book_confirm_delete.html'):
    book= get_object_or_404(Book, pk=pk)    
    if request.method=='POST':
        book.delete()
        return redirect('book_list')
    return render(request, template_name, {'object':book})
